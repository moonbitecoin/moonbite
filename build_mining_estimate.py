"""Build BigCoin mining-time estimate workbook (Litecoin/Scrypt fork).

All results are live formulas referencing the Assumptions sheet — change an
input and every downstream figure recalculates.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")          # hardcoded inputs
BLACK = Font(name=FONT, color="000000")         # formulas
GREEN = Font(name=FONT, color="008000")         # cross-sheet links
BOLD = Font(name=FONT, bold=True)
TITLE = Font(name=FONT, bold=True, size=14)
HDR = Font(name=FONT, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREY = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center")

wb = Workbook()

# ----------------------------------------------------------------- Assumptions
a = wb.active
a.title = "Assumptions"
a["A1"] = "BigCoin Mining Estimate — Assumptions (Litecoin / Scrypt fork)"
a["A1"].font = TITLE
a.merge_cells("A1:D1")

def hdr(ws, row, cols):
    for c, txt in cols.items():
        cell = ws[f"{c}{row}"]
        cell.value = txt
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = WRAP

hdr(a, 3, {"A": "Parameter", "B": "Value", "C": "Unit", "D": "Notes / Source"})

# Chain / emission params (label, value, unit, note)
chain_rows = [
    ("Total circulation target", 15000000, "BIG", "User-specified target to mine."),
    ("Initial block reward", 50, "BIG/block", "BigCoin genesis reward (Litecoin-style)."),
    ("Block time", 150, "seconds", "2.5 minutes — Litecoin target spacing."),
    ("Halving interval", 840000, "blocks", "Litecoin schedule (reward halves each interval)."),
    ("Scrypt difficulty-1 work", 4294967296, "hashes", "2^32 — hashes per block at difficulty 1 (Bitcoin/Litecoin definition)."),
    ("Solo mining difficulty", 1, "(unitless)", "Fresh-fork minimum. RAISE THIS to model a harder/established chain."),
    ("Assumed network hashrate", 1000000000000, "H/s", "1 TH/s — for the 'share of network' model only. Adjust to your network."),
]
r = 4
for label, val, unit, note in chain_rows:
    a[f"A{r}"] = label; a[f"A{r}"].font = BLACK
    a[f"B{r}"] = val; a[f"B{r}"].font = BLUE
    a[f"C{r}"] = unit; a[f"C{r}"].font = BLACK
    a[f"D{r}"] = note; a[f"D{r}"].font = BLACK; a[f"D{r}"].alignment = WRAP
    for c in "ABCD":
        a[f"{c}{r}"].border = BORDER
    a[f"B{r}"].fill = YELLOW
    r += 1
a["B8"].comment = Comment("2^32 = 4,294,967,296. Expected hashes to find a block = difficulty x 2^32.", "BigCoin")
a["B4"].number_format = "#,##0"
a["B7"].number_format = "#,##0"
a["B8"].number_format = "#,##0"
a["B10"].number_format = "#,##0"

# Hardware table
hw_title_row = 12
a[f"A{hw_title_row}"] = "Mining hardware (Scrypt)"
a[f"A{hw_title_row}"].font = BOLD
hdr(a, hw_title_row + 1, {"A": "Hardware", "B": "Hashrate (H/s)", "C": "Power (W)", "D": "Notes / Source"})
hw_rows = [
    ("CPU (modern desktop x86)", 50000, 95, "~50 KH/s Scrypt; typical multi-core desktop CPU."),
    ("GPU (high-end)", 750000, 200, "~0.75 MH/s Scrypt; modern high-end GPU."),
    ("ASIC (Bitmain Antminer L7)", 9500000000, 3425, "9.5 GH/s @ 3425 W — Bitmain L7 published spec."),
]
hw_first = hw_title_row + 2  # row 14
r = hw_first
for label, hr, pw, note in hw_rows:
    a[f"A{r}"] = label; a[f"A{r}"].font = BLACK
    a[f"B{r}"] = hr; a[f"B{r}"].font = BLUE; a[f"B{r}"].number_format = "#,##0"
    a[f"C{r}"] = pw; a[f"C{r}"].font = BLUE; a[f"C{r}"].number_format = "#,##0"
    a[f"D{r}"] = note; a[f"D{r}"].font = BLACK; a[f"D{r}"].alignment = WRAP
    for c in "ABCD":
        a[f"{c}{r}"].border = BORDER
    a[f"B{r}"].fill = YELLOW; a[f"C{r}"].fill = YELLOW
    r += 1
a[f"B{hw_first+2}"].comment = Comment("Antminer L7: 9,500 MH/s = 9.5 GH/s, ~3425W. Source: Bitmain spec sheet.", "BigCoin")

hw_cpu, hw_gpu, hw_asic = hw_first, hw_first + 1, hw_first + 2  # 14,15,16

# Legend
leg = hw_first + 4
a[f"A{leg}"] = "Legend:"; a[f"A{leg}"].font = BOLD
a[f"A{leg+1}"] = "Yellow / blue text = input you can edit."; a[f"A{leg+1}"].font = BLUE
a[f"A{leg+2}"] = "Black = formula.  Green = value pulled from another sheet."; a[f"A{leg+2}"].font = BLACK

for col, w in {"A": 30, "B": 20, "C": 14, "D": 52}.items():
    a.column_dimensions[col].width = w

# ------------------------------------------------------------- Emission Schedule
e = wb.create_sheet("Emission Schedule")
e["A1"] = "Emission Schedule & Blocks to Reach Target"
e["A1"].font = TITLE
e.merge_cells("A1:H1")
e["A2"] = ("Reward halves every interval. 15M is reached in Epoch 0 (before the first "
           "halving at 42M emitted), so reward stays 50 BIG the whole way.")
e["A2"].font = Font(name=FONT, italic=True); e.merge_cells("A2:H2"); e["A2"].alignment = WRAP

cols = {"A": "Epoch", "B": "Block Reward (BIG)", "C": "Halving Interval (blocks)",
        "D": "Coins in Epoch (BIG)", "E": "Cumulative Before (BIG)",
        "F": "Cumulative After (BIG)", "G": "Remaining to Target Before (BIG)",
        "H": "Blocks Toward Target"}
hdr(e, 4, cols)

first = 5
n_epochs = 10
for i in range(n_epochs):
    row = first + i
    e[f"A{row}"] = i
    e[f"B{row}"] = f"=Assumptions!$B$5/2^A{row}"
    e[f"C{row}"] = "=Assumptions!$B$7"
    e[f"D{row}"] = f"=B{row}*C{row}"
    if i == 0:
        e[f"E{row}"] = 0
    else:
        e[f"E{row}"] = f"=F{row-1}"
    e[f"F{row}"] = f"=E{row}+D{row}"
    e[f"G{row}"] = f"=MAX(0,Assumptions!$B$4-E{row})"
    e[f"H{row}"] = f"=IF(G{row}<=0,0,MIN(C{row},ROUNDUP(G{row}/B{row},0)))"
    for c in "ABCDEFGH":
        e[f"{c}{row}"].border = BORDER
        e[f"{c}{row}"].font = BLACK
    for c in "BDEFG":
        e[f"{c}{row}"].number_format = "#,##0"
    e[f"H{row}"].number_format = "#,##0"

last = first + n_epochs - 1
tot = last + 1
e[f"A{tot}"] = "TOTAL"; e[f"A{tot}"].font = BOLD
e[f"H{tot}"] = f"=SUM(H{first}:H{last})"; e[f"H{tot}"].font = BOLD
e[f"H{tot}"].number_format = "#,##0"; e[f"H{tot}"].fill = GREY

# Derived outputs
d0 = tot + 2
outputs = [
    ("Blocks to reach target", f"=H{tot}", "blocks", "#,##0"),
    ("Block time", "=Assumptions!$B$6", "seconds", "#,##0"),
    ("Calendar time to target (live network)", f"=B{d0}*B{d0+1}", "seconds", "#,##0"),
    ("  in days", f"=B{d0+2}/86400", "days", "#,##0.0"),
    ("  in years", f"=B{d0+3}/365.25", "years", "#,##0.00"),
]
r = d0
for label, formula, unit, fmt in outputs:
    e[f"A{r}"] = label; e[f"A{r}"].font = BOLD if r in (d0, d0+2) else BLACK
    e[f"B{r}"] = formula; e[f"B{r}"].font = BLACK; e[f"B{r}"].number_format = fmt
    e[f"C{r}"] = unit; e[f"C{r}"].font = BLACK
    r += 1
e[f"A{d0}"].fill = GREY; e[f"B{d0}"].fill = GREY
BLOCKS_REF = f"'Emission Schedule'!$B${d0}"          # blocks to target
CAL_DAYS_REF = f"'Emission Schedule'!$B${d0+3}"      # calendar days (network)

for col, w in {"A": 34, "B": 20, "C": 22, "D": 20, "E": 20, "F": 20, "G": 24, "H": 18}.items():
    e.column_dimensions[col].width = w

# ------------------------------------------------------------ Hardware Comparison
h = wb.create_sheet("Hardware Comparison")
h["A1"] = "CPU vs GPU vs ASIC — Time to Mine 15,000,000 BIG"
h["A1"].font = TITLE
h.merge_cells("A1:I1")

# --- Model A: solo at fixed difficulty ---
h["A3"] = "MODEL A — Solo mining at FIXED difficulty (brand-new / isolated chain, hashrate-bound)"
h["A3"].font = BOLD; h.merge_cells("A3:I3")
h["A4"] = ("Time = Blocks to target x Difficulty x 2^32 / Hashrate. This is where hardware matters. "
           "Uses 'Solo mining difficulty' from Assumptions (default 1 = fresh fork).")
h["A4"].font = Font(name=FONT, italic=True); h.merge_cells("A4:I4"); h["A4"].alignment = WRAP

acols = {"A": "Hardware", "B": "Hashrate (H/s)", "C": "Speed vs CPU (x)",
         "D": "Total Hashes Needed", "E": "Time (seconds)", "F": "Time (days)",
         "G": "Time (years)", "H": "Power (W)", "I": "Energy (kWh)"}
hdr(h, 5, acols)

hw_src = [("CPU", hw_cpu), ("GPU", hw_gpu), ("ASIC", hw_asic)]
arow0 = 6
for idx, (name, srow) in enumerate(hw_src):
    row = arow0 + idx
    h[f"A{row}"] = name
    h[f"B{row}"] = f"=Assumptions!$B${srow}"; h[f"B{row}"].font = GREEN
    h[f"C{row}"] = f"=B{row}/$B${arow0}"
    h[f"D{row}"] = f"={BLOCKS_REF}*Assumptions!$B$9*Assumptions!$B$8"
    h[f"E{row}"] = f"=D{row}/B{row}"
    h[f"F{row}"] = f"=E{row}/86400"
    h[f"G{row}"] = f"=F{row}/365.25"
    h[f"H{row}"] = f"=Assumptions!$C${srow}"; h[f"H{row}"].font = GREEN
    h[f"I{row}"] = f"=H{row}*E{row}/3600/1000"
    for c in "ABCDEFGHI":
        h[f"{c}{row}"].border = BORDER
        if h[f"{c}{row}"].font.color is None or c in "ACDEFGI":
            h[f"{c}{row}"].font = BLACK if c in "ACDEFGI" else h[f"{c}{row}"].font
    h[f"B{row}"].number_format = "#,##0"
    h[f"C{row}"].number_format = "#,##0"
    h[f"D{row}"].number_format = "0.000E+00"
    h[f"E{row}"].number_format = "#,##0"
    h[f"F{row}"].number_format = "#,##0.0"
    h[f"G{row}"].number_format = "#,##0.00"
    h[f"H{row}"].number_format = "#,##0"
    h[f"I{row}"].number_format = "#,##0"
# re-assert green links (loop above may have overwritten)
for idx, (name, srow) in enumerate(hw_src):
    row = arow0 + idx
    h[f"B{row}"].font = GREEN
    h[f"H{row}"].font = GREEN

# --- Model B: live network share ---
b0 = arow0 + 4  # row 10
h[f"A{b0}"] = "MODEL B — Live network (difficulty retargets to hold 2.5-min blocks)"
h[f"A{b0}"].font = BOLD; h.merge_cells(f"A{b0}:I{b0}")
h[f"A{b0+1}"] = ("Calendar time to emit 15M is FIXED by the protocol and is the SAME for all hardware "
                 "(see cell below). Hardware only changes YOUR share of the rewards.")
h[f"A{b0+1}"].font = Font(name=FONT, italic=True); h.merge_cells(f"A{b0+1}:I{b0+1}")
h[f"A{b0+1}"].alignment = WRAP

h[f"A{b0+2}"] = "Calendar time to mine 15M on live network (all hardware):"
h[f"A{b0+2}"].font = BOLD; h.merge_cells(f"A{b0+2}:E{b0+2}")
h[f"F{b0+2}"] = f"={CAL_DAYS_REF}"; h[f"F{b0+2}"].font = GREEN
h[f"F{b0+2}"].number_format = "#,##0.0"; h[f"F{b0+2}"].fill = GREY
h[f"G{b0+2}"] = "days"; h[f"G{b0+2}"].font = BLACK

bcols = {"A": "Hardware", "B": "Hashrate (H/s)", "C": "Network Hashrate (H/s)",
         "D": "Your Share", "E": "Expected Coins / Day", "F": "Expected Days per Block (50 BIG)",
         "G": "Expected Coins over the run"}
hdr(h, b0 + 3, bcols)
brow0 = b0 + 4
daily_emission = "(86400/Assumptions!$B$6)*Assumptions!$B$5"
for idx, (name, srow) in enumerate(hw_src):
    row = brow0 + idx
    h[f"A{row}"] = name; h[f"A{row}"].font = BLACK
    h[f"B{row}"] = f"=Assumptions!$B${srow}"; h[f"B{row}"].font = GREEN; h[f"B{row}"].number_format = "#,##0"
    h[f"C{row}"] = "=Assumptions!$B$10"; h[f"C{row}"].font = GREEN; h[f"C{row}"].number_format = "#,##0"
    h[f"D{row}"] = f"=B{row}/C{row}"; h[f"D{row}"].font = BLACK; h[f"D{row}"].number_format = "0.0000%"
    h[f"E{row}"] = f"=D{row}*({daily_emission})"; h[f"E{row}"].font = BLACK; h[f"E{row}"].number_format = "#,##0.00"
    h[f"F{row}"] = f"=(Assumptions!$B$6/D{row})/86400"; h[f"F{row}"].font = BLACK; h[f"F{row}"].number_format = "#,##0.000"
    h[f"G{row}"] = f"=D{row}*Assumptions!$B$4"; h[f"G{row}"].font = BLACK; h[f"G{row}"].number_format = "#,##0"
    for c in "ABCDEFG":
        h[f"{c}{row}"].border = BORDER

for col, w in {"A": 26, "B": 18, "C": 20, "D": 12, "E": 18, "F": 24, "G": 22, "H": 12, "I": 14}.items():
    h.column_dimensions[col].width = w

# ------------------------------------------------------------------- Summary (1st)
s = wb.create_sheet("Summary", 0)
s["A1"] = "BigCoin — Time to Mine 15,000,000 Coins (Litecoin/Scrypt fork)"
s["A1"].font = TITLE; s.merge_cells("A1:D1")
s["A3"] = ("Two honest models. On a LIVE network, difficulty retargeting fixes the calendar time "
           "regardless of hardware — so the CPU/GPU/ASIC race only exists when you solo-mine an "
           "isolated chain at a fixed low difficulty (Model A), or it shows up as your SHARE of a "
           "live network (Model B).")
s["A3"].font = Font(name=FONT, italic=True); s.merge_cells("A3:D5"); s["A3"].alignment = WRAP

s["A7"] = "Headline figures"; s["A7"].font = BOLD
hdr(s, 8, {"A": "Metric", "B": "Value", "C": "Unit", "D": ""})
summ = [
    ("Blocks to reach 15,000,000 BIG", f"={BLOCKS_REF}", "blocks", "#,##0"),
    ("Live-network calendar time (any hardware)", f"={CAL_DAYS_REF}", "days", "#,##0.0"),
    ("Solo @ diff-1 — CPU", "='Hardware Comparison'!G6", "years", "#,##0.00"),
    ("Solo @ diff-1 — GPU", "='Hardware Comparison'!G7", "years", "#,##0.00"),
    ("Solo @ diff-1 — ASIC (L7)", "='Hardware Comparison'!F8", "days", "#,##0.00"),
]
r = 9
for label, formula, unit, fmt in summ:
    s[f"A{r}"] = label; s[f"A{r}"].font = BLACK
    s[f"B{r}"] = formula; s[f"B{r}"].font = GREEN; s[f"B{r}"].number_format = fmt; s[f"B{r}"].fill = GREY
    s[f"C{r}"] = unit; s[f"C{r}"].font = BLACK
    for c in "ABC":
        s[f"{c}{r}"].border = BORDER
    r += 1
s["A16"] = "See 'Assumptions' to change inputs; 'Emission Schedule' and 'Hardware Comparison' for the math."
s["A16"].font = Font(name=FONT, italic=True); s.merge_cells("A16:D16")
for col, w in {"A": 42, "B": 18, "C": 10, "D": 6}.items():
    s.column_dimensions[col].width = w

# default font pass for any stragglers
for ws in wb.worksheets:
    for rowc in ws.iter_rows():
        for cell in rowc:
            if cell.value is not None and cell.font.name != FONT:
                cell.font = Font(name=FONT, bold=cell.font.bold, italic=cell.font.italic,
                                 color=cell.font.color, size=cell.font.size)

# Force Excel to recalculate every formula on open (no cached values are
# written by openpyxl, so this guarantees correct figures without LibreOffice).
wb.calculation.fullCalcOnLoad = True

out = r"C:\Users\usman\Desktop\BigCoinBB\BigCoin_Mining_Estimate.xlsx"
wb.save(out)
print("saved", out)
